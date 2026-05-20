"""Deep dive: ENE rows 60-200 + special sheets to understand the CEO financial model."""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl


def dump_range(ws, row_from: int, row_to: int, col_from: int = 1, col_to: int = 20) -> None:
    print(f"  range: R{row_from}-R{row_to}, C{col_from}-C{col_to}")
    for r in range(row_from, row_to + 1):
        cells = []
        for c in range(col_from, col_to + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                cells.append("")
            else:
                s = str(v)
                if len(s) > 50:
                    s = s[:47] + "..."
                cells.append(s)
        line = " | ".join(cells).rstrip(" |")
        if line.strip():
            print(f"  R{r:>3}: {line}")


def main() -> None:
    path = sys.argv[1] if len(sys.argv) > 1 else r"C:/Users/Santiago/Downloads/Presupuestos Empresa.xlsx"
    wb = openpyxl.load_workbook(path, data_only=False, read_only=False)

    print("=" * 80)
    print("DEEP DIVE: ENE detail (gastos block)")
    print("=" * 80)
    ene = wb["ENE"]
    print("ENE rows 60-200, cols A-T")
    dump_range(ene, 60, 200, 1, 20)

    print("\n" + "=" * 80)
    print("DEEP DIVE: ENE rows 60-200, cols U-AT (deudas, ahorros, sub-categorías)")
    print("=" * 80)
    dump_range(ene, 60, 200, 21, 46)

    special = ["MEJORAR MI PRESUPUESTO", "DISTRIBUCION", "FONDO DE AMORTIZACION",
               "CALCULADORA DE DEUDAS", "PATRIMONIO NETO", "CALENDARIO", "INVERSIONES", "RETO"]
    for name in special:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        print("\n" + "=" * 80)
        print(f"DEEP DIVE: {name}  dims={ws.dimensions}  max_row={ws.max_row} max_col={ws.max_column}")
        print("=" * 80)
        dump_range(ws, 1, min(ws.max_row, 80), 1, min(ws.max_column, 22))


if __name__ == "__main__":
    main()
