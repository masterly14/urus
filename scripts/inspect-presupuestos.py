"""Inspect the CEO's budget Excel to map its semantics into the new Finanzas module."""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl


def main(path: str) -> None:
    wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
    print(f"FILE: {path}")
    print(f"SHEETS ({len(wb.sheetnames)}): {wb.sheetnames}")
    print("=" * 80)

    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n--- SHEET: {name} ---")
        print(f"  dims: {ws.dimensions}  max_row={ws.max_row}  max_col={ws.max_column}")
        print(f"  merged: {sorted([str(r) for r in ws.merged_cells.ranges])[:10]}")

        rows_to_show = min(ws.max_row, 60)
        cols_to_show = min(ws.max_column, 18)
        for r in range(1, rows_to_show + 1):
            cells = []
            for c in range(1, cols_to_show + 1):
                cell = ws.cell(row=r, column=c)
                v = cell.value
                if v is None:
                    cells.append("")
                else:
                    s = str(v)
                    if len(s) > 40:
                        s = s[:37] + "..."
                    cells.append(s)
            line = " | ".join(cells).rstrip(" |")
            if line.strip():
                print(f"  R{r:>3}: {line}")
        if ws.max_row > rows_to_show:
            print(f"  ... ({ws.max_row - rows_to_show} más filas)")


if __name__ == "__main__":
    target = sys.argv[1] if len(sys.argv) > 1 else r"C:/Users/Santiago/Downloads/Presupuestos Empresa.xlsx"
    main(target)
